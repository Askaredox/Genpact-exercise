import time
import re
import shutil
import openpyxl as xl
from os import path as ospath, makedirs
from watchdog.observers import Observer
from watchdog.events import LoggingEventHandler


class Watcher:
    """The Watcher class provides a lookup of the observed folder to make the process"""
    observer = None
    lookup_path = None
    processed_path = 'processed'
    noapplicable_path = 'no_applicable'
    master_workbook_path = 'master_workbook.xlsx'
    excel_format = r'^.+\.xlsx?$'

    def __init__(self):
        self.event_handler = LoggingEventHandler()

    def __check_dirs(self):
        """Check if the directories exists, if not they are created"""
        paths = [
            self.lookup_path,
            self.processed_path,
            self.noapplicable_path
        ]
        for path in paths:
            if not ospath.exists(path):
                makedirs(path)

        if not ospath.exists(self.master_workbook_path):
            wb = xl.Workbook()
            wb.save(self.master_workbook_path)

    def __on_created(self, event):
        """If a file is created or moved into the lookup_path, are processed to join into a master workbook
            if not, then the file is moved into the noapplicable_path.
            For the use of the observer.on_created event_handler
        """
        time.sleep(.5)
        if (event.is_directory):
            return
        path = event.src_path
        move_dir = self.noapplicable_path
        if (self.__is_excel_file(path)):
            move_dir = self.processed_path
            self.__copy_sheets(path)

        shutil.move(path, move_dir)

    def __copy_sheets(self, src:str):
        """Copy the contents of the workbook into the master workbook"""
        master = xl.load_workbook(filename=self.master_workbook_path)
        wb = xl.load_workbook(filename=src)

        for sheet in wb.worksheets:
            ws = master.create_sheet()
            for row in sheet:
                for cell in row:
                    ws[cell.coordinate].value = cell.value
        master.save(filename=self.master_workbook_path)

    def __is_excel_file(self, path:str):
        """checks if the file on path is an Excel file by a regex match"""
        _, file = ospath.split(path)
        match = re.search(self.excel_format, file)
        return match is not None

    def observe(self, path:str):
        """Start observing the folder for a new file created or moved
        
        + path(str): path of the directory to be observed"""
        self.lookup_path = path
        self.__check_dirs()
        self.observer = Observer()
        self.observer.schedule(self.event_handler, self.lookup_path, recursive=True)
        self.event_handler.on_created = self.__on_created
        self.observer.start()
        #try:
        #    while (True):
        #        time.sleep(1)
        #except KeyboardInterrupt:
        #    self.observer.stop()
        #self.observer.join()
    
    def pause(self):
        """Pause for a moment the observing state"""
        self.observer.stop()

    def stop(self):
        """completly stop the class to observe the folder"""
        self.observer.join()


