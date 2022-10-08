import logging
import time
import re
import shutil
import openpyxl as xl
from os import path as ospath, makedirs
from watchdog.observers import Observer
from watchdog.events import LoggingEventHandler


class Watcher:
    observer = None
    lookup_path = None
    processed_path = 'processed'
    noapplicable_path = 'no_applicable'
    master_workbook_path = 'master_workbook.xlsx'
    excel_format = r'^.+\.xlsx?$'

    def __init__(self, path):
        self.lookup_path = path
        self.check_dirs()
        event_handler = LoggingEventHandler()
        self.observer = Observer()
        self.observer.schedule(event_handler, self.lookup_path, recursive=True)
        event_handler.on_created = self.on_created

    def check_dirs(self):
        paths = [
            self.lookup_path,
            self.processed_path,
            self.noapplicable_path
        ]
        for path in paths:
            if not ospath.exists(path):
                makedirs(path)

        if not ospath.exists(self.master_workbook_path):
            wb = xl.Workbook()
            wb.save(self.master_workbook_path)

    def on_created(self, event):
        time.sleep(.5)
        if (event.is_directory):
            return
        path = event.src_path
        move_dir = self.noapplicable_path
        if (self.is_excel_file(path)):
            move_dir = self.processed_path
            self.copy_sheets(path)

        shutil.move(path, move_dir)

    def copy_sheets(self, src):
        master = xl.load_workbook(filename=self.master_workbook_path)
        wb = xl.load_workbook(filename=src)

        for sheet in wb.worksheets:
            ws = master.create_sheet()
            for row in sheet:
                for cell in row:
                    ws[cell.coordinate].value = cell.value
        master.save(filename=self.master_workbook_path)

    def is_excel_file(self, path):
        _, file = ospath.split(path)
        match = re.search(self.excel_format, file)
        return match is not None

    def observe(self):
        self.observer.start()
        try:
            while (True):
                time.sleep(1)
        except KeyboardInterrupt:
            self.observer.stop()
        self.observer.join()


def main():
    logging.basicConfig(
        level=logging.INFO, format='%(asctime)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    path = input('Input the lookup folder ["./lookup"]:')
    w = Watcher(path if path != '' else 'lookup')
    w.observe()


if __name__ == "__main__":
    main()
